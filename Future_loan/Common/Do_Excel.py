from openpyxl import load_workbook
from Future_loan.Common import Project_Path
from Future_loan.Common.GetConf import GetConf
from Future_loan.Common.Test_log import TestLog
from Future_loan.Common.Mysql import Mysql


class Do_Excel:
    '''在Excel中读取测试数据、写回测试结果'''

    def __init__(self):
        # global filename#声明为全局属性
        # self.filename=GetConf('WORKBOOK','filename').get_Str()#根据配置文件获取工作簿名称
        # # global sheetname#声明为全局属性
        # self.sheetname=GetConf('WORKBOOK','sheetname').get_Str()#根据配置文件获取表单名
        self.log=TestLog()
        self.mode=GetConf('CASEID','mode').get_Int()

    def read_Excel(self,sheetname):
        '''根据指定mode读取数据
        0：读取所有数据
        1：读取指定id的数据
        2：读取指定模块的数据'''

        read=load_workbook(Project_Path.testcase_path)#打开文件
        sheet=read[sheetname]#定位表单


        all_data=[]
        # row_data = {}
        tel=self.get_tel()#获取电话号码
        try:
            if self.mode ==0:#当mode为0时，读取所有数据
                for row in range(2,sheet.max_row+1):#从第2行开始读取数据
                        row_data = {}
                        row_data['Case_Id'] = sheet.cell(row, 1).value
                        row_data['Module'] = sheet.cell(row, 2).value
                        row_data['Url'] = sheet.cell(row, 3).value
                        row_data['Description'] = sheet.cell(row, 4).value
                        row_data['Method'] = sheet.cell(row, 5).value
                        row_data['Params'] = sheet.cell(row, 6).value
                        if 'tel' in row_data['Params']:
                            row_data['Params']=sheet.cell(row,6).value.replace('tel',str(tel))
                            self.write_excel('tel',1,2,int(tel)+1)
                        # if sheet.cell(row,7).value!=None:
                        row_data['Sql']=sheet.cell(row,7).value
                        row_data['ExpectedResult'] = sheet.cell(row, 8).value
                        all_data.append(row_data)
            elif self.mode==1:#当mode为1时，读取指定id的数据
                case_id=GetConf('CASEID','id').get_Tuple_list()
                for row in case_id:  # 遍历获取到caseid列表，根据caseid读取响应数据
                    row_data = {}
                    row_data['Case_Id'] = sheet.cell(row+1, 1).value  #
                    row_data['Module'] = sheet.cell(row+1, 2).value
                    row_data['Url'] = sheet.cell(row+1, 3).value
                    row_data['Description'] = sheet.cell(row+1, 4).value
                    row_data['Method'] = sheet.cell(row+1, 5).value
                    row_data['Params'] = sheet.cell(row+1, 6).value
                    if 'tel' in row_data['Params']:
                        row_data['Params'] = sheet.cell(row, 6).value.replace('tel', str(tel))
                        self.write_excel('tel', 1, 2, int(tel) + 1)
                    row_data['Sql'] = sheet.cell(row, 7).value
                    row_data['ExpectedResult'] = sheet.cell(row+1, 8).value
                    all_data.append(row_data)

            elif self.mode==2:#当mode为2时，读取指定模块的数据
                modul=GetConf('CASEID','module').get_Str()
                for row in range(2,sheet.max_row+1):#从第2行开始读取数据
                        row_data = {}
                        row_data['Case_Id'] = sheet.cell(row, 1).value
                        row_data['Module'] = sheet.cell(row, 2).value
                        row_data['Url'] = sheet.cell(row, 3).value
                        row_data['Description'] = sheet.cell(row, 4).value
                        row_data['Method'] = sheet.cell(row, 5).value
                        row_data['Params'] = sheet.cell(row, 6).value
                        if 'tel' in row_data['Params']:
                            row_data['Params']=sheet.cell(row,6).value.replace('tel',str(tel))
                            self.write_excel('tel',1,2,int(tel)+1)
                        row_data['Sql'] = sheet.cell(row, 7).value
                        row_data['ExpectedResult'] = sheet.cell(row, 8).value
                        if row_data['Module']==modul:
                            all_data.append(row_data)


        except Exception as e:
            self.log.error('数据读取出错：{}'.format(e))

        read.close()

        return all_data

    # def read_excel(self,sheetname):
    #     '''读取数据'''
    #     read=load_workbook(Project_Path.testcase_path)#打开文件
    #     sheet=read[sheetname]#定位表单
    #
    #
    #     all_data=[]
    #     # row_data = {}
    #     tel=self.get_tel()#获取电话号码
    #     try:
    #         if self.mode ==0:
    #             for row in range(2,sheet.max_row+1):#从第2行开始读取数据
    #                     row_data = {}
    #                     row_data['Case_Id'] = sheet.cell(row, 1).value
    #                     row_data['Module'] = sheet.cell(row, 2).value
    #                     row_data['Url'] = sheet.cell(row, 3).value
    #                     row_data['Description'] = sheet.cell(row, 4).value
    #                     row_data['Method'] = sheet.cell(row, 5).value
    #                     row_data['Params'] = sheet.cell(row, 6).value
    #                     if 'tel' in row_data['Params']:
    #                         row_data['Params']=sheet.cell(row,6).value.replace('tel',str(tel))
    #                         self.write_excel('tel',1,2,int(tel)+1)
    #
    #                     row_data['ExpectedResult'] = sheet.cell(row, 7).value
    #                     all_data.append(row_data)
    #         elif self.mode==1:
    #             case_id=GetConf('CASEID','id').get_Tuple_list()
    #             for row in case_id:  # 遍历获取到caseid列表，根据caseid读取响应数据
    #                 row_data = {}
    #                 row_data['Case_Id'] = sheet.cell(row+1, 1).value  #
    #                 row_data['Module'] = sheet.cell(row+1, 2).value
    #                 row_data['Url'] = sheet.cell(row+1, 3).value
    #                 row_data['Description'] = sheet.cell(row+1, 4).value
    #                 row_data['Method'] = sheet.cell(row+1, 5).value
    #                 row_data['Params'] = sheet.cell(row+1, 6).value
    #                 row_data['ExpectedResult'] = sheet.cell(row+1, 7).value
    #                 all_data.append(row_data)
    #
    #     except Exception as e:
    #         self.log.error('数据读取出错：{}'.format(e))
    #
    #     read.close()
    #
    #     return all_data


    def write_excel(self,sheetname,row,column,data):
        '''写入数据'''
        write=load_workbook(Project_Path.testcase_path)#打开文件
        sheet=write[sheetname]#表单
        try:
            # if data =='Pass':

                sheet.cell(row=row,column=column,value=data)#写入数据
            # else:
            #
            #     sheet.cell(row=row, column=column,value=data,style=font)  # 写入数据
        except Exception as e:
            self.log.error('数据写入出错：{}'.format(e))

        write.save(Project_Path.testcase_path)#保存
        write.close()


    def get_tel(self):
        '''获取‘tel’表单里的电话号码'''
        read = load_workbook(Project_Path.testcase_path)  # 打开文件
        sheet = read['tel']  # 定位表单
        tel=sheet.cell(1,2).value
        read.close()
        return tel


if __name__ == '__main__':
    # case_id = GetConf('TestCaseID.conf','CASEID','id_list').get_List_Tuple_Dict()
    res=Do_Excel().read_Excel('recharge')
    print(res)

